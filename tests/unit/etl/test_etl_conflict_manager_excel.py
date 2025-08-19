import os
import pytest
import pandas as pd
from openpyxl import load_workbook
from db.models.model_curation import (
    CurationConflict,
    ConflictStatus,
    ConflictResolution,
)
from biofilter.etl.conflict_manager import ConflictManager


@pytest.fixture
def conflict_manager(db_session):
    from biofilter.utils.logger import Logger

    return ConflictManager(session=db_session, logger=Logger())


@pytest.fixture
def sample_conflicts(db_session):
    conflict_1 = CurationConflict(
        entity_type="gene",
        identifier="HGNC:1234",
        existing_identifier="HGNC:5678",
        status=ConflictStatus.pending,
        resolution=None,
        description="Test conflict A",
        entity_id=1,
    )
    conflict_2 = CurationConflict(
        entity_type="gene",
        identifier="HGNC:9999",
        existing_identifier="HGNC:8888",
        status=ConflictStatus.pending,
        resolution=None,
        description="Test conflict B",
        entity_id=2,
    )
    db_session.add_all([conflict_1, conflict_2])
    db_session.commit()
    return [conflict_1, conflict_2]


def test_export_conflicts_to_excel(
    tmp_path, db_session, conflict_manager, sample_conflicts
):  # noqa E501
    output_file = tmp_path / "conflicts.xlsx"
    success = conflict_manager.export_conflicts_to_excel(str(output_file))
    assert success
    assert os.path.exists(output_file)

    wb = load_workbook(output_file)
    ws = wb["Conflicts"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][:5] == (
        "id",
        "entity_type",
        "entity_id",
        "identifier",
        "existing_identifier",
    )
    assert len(rows) == 3  # 2 conflicts + 1 header


def test_import_conflicts_from_excel(
    tmp_path, db_session, conflict_manager, sample_conflicts
):  # noqa E501
    # Export first to generate file
    output_file = tmp_path / "conflicts_import.xlsx"
    conflict_manager.export_conflicts_to_excel(str(output_file))

    # Modify Excel manually via Pandas for the test
    df = pd.read_excel(output_file)
    df.loc[0, "status"] = "resolved"
    df.loc[0, "resolution"] = "merge"
    df.loc[0, "notes"] = "Merged successfully"
    df.to_excel(output_file, index=False)

    updated = conflict_manager.import_conflicts_from_excel(str(output_file))
    assert updated == 1

    updated_conflict = (
        db_session.query(CurationConflict).filter_by(id=1).first()
    )  # noqa E501
    assert updated_conflict.status == ConflictStatus.resolved
    assert updated_conflict.resolution == ConflictResolution.merge
    assert updated_conflict.notes == "Merged successfully"
